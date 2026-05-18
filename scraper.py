#!/usr/bin/env python3
"""
PiP Studio - Buscador de Leads
- Multiples queries por categoria para maximizar resultados
- Email scraping (home + pagina de contacto)
- Normalizacion de telefonos al formato argentino
- Columna "Tiene web" (Si/No)
- Progreso via callback para SSE
"""

from __future__ import annotations

import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------
GOOGLE_API_KEY        = "AIzaSyD1fbjrqYM7kfGEt-LRAlYFl_IvMP-Qrps"
TEXT_SEARCH_URL       = "https://places.googleapis.com/v1/places:searchText"
PAGE_SIZE             = 20
MAX_PAGES             = 3
REGION_CODE           = "AR"
LANGUAGE_CODE         = "es"
DELAY_BETWEEN_QUERIES = 0.5
EMAIL_TIMEOUT         = 6
MAX_CITIES            = 5

OUTPUT_DIR = Path(__file__).resolve().parent / "output"

FIELD_MASK = ",".join([
    "places.displayName",
    "places.formattedAddress",
    "places.nationalPhoneNumber",
    "places.internationalPhoneNumber",
    "places.websiteUri",
    "places.addressComponents",
    "nextPageToken",
])

EXCEL_HEADERS = ["Nombre", "Telefono", "Email", "Direccion", "Ciudad", "Tiene web", "Sitio web"]

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE)

EMAIL_BLACKLIST = {
    "example.com", "sentry.io", "wixpress.com", "squarespace.com",
    "wordpress.com", "shopify.com", "googleapis.com", "gstatic.com",
    "facebook.com", "instagram.com", "twitter.com", "youtube.com",
    "w3.org", "schema.org", "jquery.com", "cloudflare.com",
    "apple.com", "microsoft.com", "adobe.com", "amazon.com",
}

MAX_PAGES_TO_CRAWL = 8   # maximo de paginas internas a visitar por sitio

PHONE_CLEAN_RE = re.compile(r"[^\d+]")


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class Category:
    label: str
    queries: tuple
    place_type: str = None


@dataclass
class ScrapeResult:
    filename: str
    path: Path
    count: int
    rubro: str
    ciudad: str
    emails_found: int = 0
    total_reported: int = None


# ---------------------------------------------------------------------------
# Categorias
# ---------------------------------------------------------------------------

CATEGORIES = (
    Category("Ferreteria", place_type="hardware_store", queries=(
        "ferreteria", "ferreteria industrial", "ferreteria y materiales",
        "correlon ferreteria", "venta de herramientas",
    )),
    Category("Veterinaria", place_type="veterinary_care", queries=(
        "veterinaria", "clinica veterinaria", "veterinario",
        "hospital veterinario", "veterinaria rural", "medico veterinario",
    )),
    Category("Agropecuaria", queries=(
        "agropecuaria", "casa agropecuaria", "venta de insumos agropecuarios",
        "agronomia", "semillas y agroquimicos", "distribuidora agropecuaria",
    )),
    Category("Cooperativa agricola", queries=(
        "cooperativa agropecuaria", "cooperativa agricola", "cooperativa rural",
        "cooperativa ganadera", "acopio de granos", "silo acopio",
    )),
    Category("Materiales de construccion", queries=(
        "correlon de materiales", "materiales de construccion", "correlon",
        "venta de materiales", "distribuidora de materiales", "hormigon y materiales",
    )),
    Category("Taller mecanico", place_type="car_repair", queries=(
        "taller mecanico", "mecanico automotor", "taller automotriz",
        "reparacion de autos", "gomeria taller", "taller de chapa y pintura",
    )),
    Category("Electricidad", place_type="electrician", queries=(
        "electricista", "electricidad", "materiales electricos",
        "instalaciones electricas", "distribuidora electrica", "iluminacion y electricidad",
    )),
    Category("Plomeria", place_type="plumber", queries=(
        "plomeria", "plomero", "sanitarios y plomeria",
        "instalaciones sanitarias", "griferias y sanitarios",
    )),
    Category("Supermercado / Mayorista", place_type="supermarket", queries=(
        "supermercado", "mayorista", "autoservicio",
        "distribuidora de alimentos", "almacen mayorista",
    )),
    Category("Farmacia", place_type="pharmacy", queries=(
        "farmacia", "drogeria", "farmacia de turno", "farmacia y perfumeria",
    )),
    Category("Metalurgica", queries=(
        "metalurgica", "torneria", "herreria", "taller metalurgico",
        "soldadura industrial", "fabricacion metalica",
    )),
    Category("Transporte / Logistica", place_type="moving_company", queries=(
        "transporte de cargas", "logistica", "fletero",
        "empresa de transporte", "transporte rural",
    )),
    Category("Ganaderia / Tambo", queries=(
        "tambo", "establecimiento ganadero", "cabana bovina",
        "feed lot", "produccion lechera",
    )),
    Category("Semillas / Agroquimicos", queries=(
        "semillas", "agroquimicos", "venta de semillas",
        "distribuidora de agroquimicos", "fitosanitarios", "fertilizantes y semillas",
    )),
    Category("Distribuidora industrial", queries=(
        "distribuidora industrial", "proveedor industrial", "insumos industriales",
        "maquinaria industrial", "repuestos industriales",
    )),
    Category("Gomeria", queries=(
        "gomeria", "neumaticos", "venta de cubiertas", "neumaticos y llantas",
    )),
    Category("Pintureria", queries=(
        "pintureria", "venta de pinturas", "pinturas y revestimientos", "pinturas industriales",
    )),
    Category("Maquinaria agricola", queries=(
        "maquinaria agricola", "concesionaria de maquinaria agricola",
        "venta de tractores", "implementos agricolas", "cosechadoras y tractores",
    )),
    Category("Dietetica / Nutricion", queries=(
        "dietetica", "nutricion", "productos naturales",
        "suplementos nutricionales", "dietetica y nutricion",
    )),
    Category("Distribuidora de bebidas", queries=(
        "distribuidora de bebidas", "venta de bebidas", "distribuidora de gaseosas",
        "distribuidora de cervezas", "bebidas y licores",
    )),
)

PROVINCE = "Santa Fe"

CITIES = (
    "Rosario", "Santa Fe", "Rafaela", "Venado Tuerto", "Reconquista",
    "Santo Tome", "Esperanza", "Casilda", "San Lorenzo", "Canada de Gomez",
    "Sunchales", "Galvez", "Vera", "Tostado", "Rufino", "Las Rosas",
    "Firmat", "Sastre", "Morteros", "San Cristobal", "Moises Ville",
    "Laguna Paiva", "Coronda", "Totoras", "Arroyo Seco", "Villa Constitucion",
    "San Justo", "Avellaneda", "Calchaqui", "Ceres",
)


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def slugify(text):
    normalized = unicodedata.normalize("NFD", text)
    without_accents = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    slug = without_accents.lower().strip()
    slug = re.sub(r"[^a-z0-9\s-]", "", slug)
    slug = re.sub(r"\s+", "-", slug)
    return re.sub(r"-+", "-", slug).strip("-")


def get_category(index):
    if index < 1 or index > len(CATEGORIES):
        raise ValueError("Elegi un numero entre 1 y {}.".format(len(CATEGORIES)))
    return CATEGORIES[index - 1]


def get_category_by_label(label):
    for cat in CATEGORIES:
        if cat.label.lower() == label.strip().lower():
            return cat
    return None


def get_city(index):
    if index < 1 or index > len(CITIES):
        raise ValueError("Elegi un numero entre 1 y {}.".format(len(CITIES)))
    return CITIES[index - 1]


def get_city_by_name(name):
    for city in CITIES:
        if city.lower() == name.strip().lower():
            return city
    return None


# ---------------------------------------------------------------------------
# Normalizacion de telefonos
# ---------------------------------------------------------------------------

def normalize_phone(raw):
    if not raw:
        return ""
    digits = PHONE_CLEAN_RE.sub("", raw)
    if digits.startswith("+54"):
        digits = digits[3:]
    elif digits.startswith("54") and len(digits) > 11:
        digits = digits[2:]
    if digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 11 and digits[2:4] == "15":
        digits = digits[:2] + digits[4:]
    elif len(digits) == 11 and digits[3:5] == "15":
        digits = digits[:3] + digits[5:]
    elif len(digits) == 11 and digits[4:6] == "15":
        digits = digits[:4] + digits[6:]
    if not digits:
        return raw.strip()
    if len(digits) == 10:
        return "+54 9 " + digits
    if 8 <= len(digits) <= 9:
        return "+54 " + digits
    return "+54 " + digits if len(digits) >= 8 else raw.strip()


# ---------------------------------------------------------------------------
# Email scraping
# ---------------------------------------------------------------------------

def is_valid_email(email):
    if not email or "@" not in email:
        return False
    domain = email.split("@")[-1].lower()
    if domain in EMAIL_BLACKLIST:
        return False
    if re.search(r"\.(png|jpg|jpeg|gif|svg|webp|ico|css|js|woff|ttf)$", domain):
        return False
    if re.search(r"\d+\.\d+", domain):
        return False
    return True


def extract_email_from_html(html):
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.startswith("mailto:"):
            email = href[7:].split("?")[0].strip().lower()
            if is_valid_email(email):
                return email
    text = soup.get_text(separator=" ")
    for match in EMAIL_REGEX.findall(text):
        email = match.strip().lower()
        if is_valid_email(email):
            return email
    return ""


def get_base_url(url):
    """Extrae el dominio base de una URL. Ej: https://miempresa.com.ar"""
    match = re.match(r"(https?://[^/]+)", url)
    return match.group(1).rstrip("/") if match else ""


def extract_internal_links(html, base_url):
    """
    Extrae todos los links internos de una pagina.
    Prioriza paginas de contacto, luego el resto.
    """
    soup  = BeautifulSoup(html, "html.parser")
    links = set()

    PRIORITY_KEYWORDS = ["contacto", "contact", "about", "nosotros", "quienes", "empresa", "info"]

    priority = []
    normal   = []

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()

        # Ignorar anchors, javascript, mailto, tel
        if not href or href.startswith("#") or href.startswith("javascript") \
                or href.startswith("mailto:") or href.startswith("tel:"):
            continue

        # Convertir a URL absoluta
        if href.startswith("http"):
            full = href
        elif href.startswith("//"):
            full = "https:" + href
        elif href.startswith("/"):
            full = base_url + href
        else:
            full = base_url + "/" + href

        # Solo links del mismo dominio
        if not full.startswith(base_url):
            continue

        # Ignorar archivos
        if re.search(r"\.(pdf|jpg|jpeg|png|gif|svg|zip|rar|docx|xlsx|mp4|mp3)$", full, re.IGNORECASE):
            continue

        # Limpiar fragmentos
        full = full.split("#")[0].rstrip("/")

        if full in links:
            continue
        links.add(full)

        # Priorizar paginas de contacto
        href_lower = href.lower()
        if any(kw in href_lower for kw in PRIORITY_KEYWORDS):
            priority.append(full)
        else:
            normal.append(full)

    return priority + normal


def scrape_email_from_url(url):
    """
    Crawlea el sitio web completo buscando un email de contacto.
    Estrategia:
    1. Visita el home
    2. Extrae todos los links internos (priorizando paginas de contacto)
    3. Visita cada pagina hasta encontrar un email o alcanzar el limite
    """
    if not url:
        return ""

    headers  = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    base_url = get_base_url(url)
    if not base_url:
        return ""

    visited      = set()
    queue_urls   = [url]
    pages_crawled = 0

    while queue_urls and pages_crawled < MAX_PAGES_TO_CRAWL:
        current_url = queue_urls.pop(0)

        if current_url in visited:
            continue
        visited.add(current_url)
        pages_crawled += 1

        try:
            r = requests.get(current_url, headers=headers, timeout=EMAIL_TIMEOUT, allow_redirects=True)
            if r.status_code != 200:
                continue

            # Buscar email en esta pagina
            email = extract_email_from_html(r.text)
            if email:
                return email

            # Si es el home, agregar links internos a la cola
            if pages_crawled == 1:
                internal_links = extract_internal_links(r.text, base_url)
                for link in internal_links:
                    if link not in visited:
                        queue_urls.append(link)

        except Exception:
            continue

    return ""


# ---------------------------------------------------------------------------
# Google Places API
# ---------------------------------------------------------------------------

def api_headers():
    return {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": GOOGLE_API_KEY,
        "X-Goog-FieldMask": FIELD_MASK,
    }


def check_api_key():
    if not GOOGLE_API_KEY or GOOGLE_API_KEY.startswith("TU_API_KEY"):
        raise ValueError("Configura GOOGLE_API_KEY al inicio de scraper.py.")


def parse_api_error(response):
    try:
        payload = response.json()
    except ValueError:
        return response.text or "HTTP {}".format(response.status_code)
    if isinstance(payload, dict):
        err = payload.get("error", {})
        if isinstance(err, dict):
            msg = err.get("message") or err.get("status")
            if msg:
                return str(msg)
    return str(payload)


def extract_city_from_components(components, fallback):
    if not components:
        return fallback
    for wanted in ("locality", "administrative_area_level_2", "administrative_area_level_1"):
        for comp in components:
            if wanted in (comp.get("types") or []):
                city = comp.get("longText") or comp.get("shortText") or ""
                if city:
                    return str(city).strip()
    return fallback


def parse_place(place, ciudad_fallback):
    dn    = place.get("displayName") or {}
    name  = dn.get("text") if isinstance(dn, dict) else str(dn)
    phone = place.get("nationalPhoneNumber") or place.get("internationalPhoneNumber") or ""
    web   = str(place.get("websiteUri") or "").strip()
    return {
        "nombre":    str(name or "").strip(),
        "telefono":  normalize_phone(str(phone).strip()),
        "email":     "",
        "direccion": str(place.get("formattedAddress") or "").strip(),
        "ciudad":    extract_city_from_components(place.get("addressComponents"), ciudad_fallback),
        "tiene_web": "Si" if web else "No",
        "sitio_web": web,
    }


def text_search_page(query, page_token=None, included_type=None):
    body = {
        "textQuery": query,
        "pageSize": PAGE_SIZE,
        "languageCode": LANGUAGE_CODE,
        "regionCode": REGION_CODE,
    }
    if included_type:
        body["includedType"] = included_type
    if page_token:
        body["pageToken"] = page_token
    response = requests.post(TEXT_SEARCH_URL, headers=api_headers(), json=body, timeout=30)
    if response.status_code != 200:
        raise requests.RequestException(parse_api_error(response))
    return response.json()


def fetch_single_query(query, included_type=None):
    places_raw = []
    page_token = None
    for page_num in range(1, MAX_PAGES + 1):
        data  = text_search_page(query, page_token, included_type=included_type if page_num == 1 else None)
        batch = data.get("places") or []
        places_raw.extend(batch)
        page_token = data.get("nextPageToken")
        if not page_token or len(places_raw) >= MAX_PAGES * PAGE_SIZE:
            break
        time.sleep(DELAY_BETWEEN_QUERIES)
    return places_raw


def deduplicate(leads):
    seen   = set()
    unique = []
    for lead in leads:
        key = (
            re.sub(r"\s+", " ", lead["nombre"].lower().strip()),
            re.sub(r"\s+", " ", lead["direccion"].lower().strip()),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(lead)
    return unique


def search_places_for_city(category, ciudad, progress=None, query_offset=0, total_queries=0):
    ciudad_label = ciudad.strip().title()
    all_raw      = []
    n            = len(category.queries)

    for i, term in enumerate(category.queries, start=1):
        full_query = "{} en {}, {}, Argentina".format(term, ciudad, PROVINCE)
        inc_type   = category.place_type if i == 1 else None

        if progress:
            current = query_offset + i
            progress('Buscando "{}" en {}...'.format(term, ciudad), current, total_queries)

        try:
            batch = fetch_single_query(full_query, included_type=inc_type)
            all_raw.extend(batch)
        except requests.RequestException:
            pass

        if i < n:
            time.sleep(DELAY_BETWEEN_QUERIES)

    leads = [parse_place(p, ciudad_label) for p in all_raw if p.get("displayName")]
    leads = [l for l in leads if l["nombre"]]
    return deduplicate(leads)


def enrich_with_emails(leads, progress=None, email_offset=0, total_emails=0):
    with_web = [l for l in leads if l.get("sitio_web")]
    if not with_web:
        return leads
    done = 0
    for lead in leads:
        if not lead.get("sitio_web"):
            continue
        done += 1
        if progress:
            progress("Buscando email: {}...".format(lead["nombre"][:40]), email_offset + done, total_emails)
        lead["email"] = scrape_email_from_url(lead["sitio_web"])
    return leads


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(leads, path):
    wb    = Workbook()
    sheet = wb.active
    sheet.title = "Leads"
    header_fill = PatternFill("solid", fgColor="1A2332")
    header_font = Font(bold=True, color="F5B800")
    sheet.append(EXCEL_HEADERS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for lead in leads:
        sheet.append([
            lead["nombre"],
            lead["telefono"],
            lead.get("email", ""),
            lead["direccion"],
            lead["ciudad"],
            lead.get("tiene_web", "No"),
            lead["sitio_web"],
        ])
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 32
    sheet.column_dimensions["D"].width = 45
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 38
    wb.save(path)


def build_output_path(category_slug, ciudad_slug, output_dir):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_dir / "leads_{}_{}_{}.xlsx".format(category_slug, ciudad_slug, ts)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def scrape_leads(category, ciudad, output_dir=None, output_file=None, verbose=True, fetch_emails=True, progress=None):
    ciudad = ciudad.strip()
    if not ciudad:
        raise ValueError("La ciudad no puede estar vacia.")
    check_api_key()

    n_queries  = len(category.queries)
    total_steps = n_queries + (50 if fetch_emails else 0)

    if verbose:
        print("\n[1/3] Buscando en Google Places - {}...".format(ciudad))

    leads = search_places_for_city(
        category, ciudad,
        progress=progress,
        query_offset=0,
        total_queries=total_steps,
    )

    with_web = sum(1 for l in leads if l.get("sitio_web"))

    if fetch_emails:
        if verbose:
            print("\n[2/3] Buscando emails en {} sitios web...".format(with_web))
        leads = enrich_with_emails(
            leads,
            progress=progress,
            email_offset=n_queries,
            total_emails=n_queries + with_web,
        )

    emails_found  = sum(1 for l in leads if l.get("email"))
    category_slug = slugify(category.label)
    ciudad_slug   = slugify(ciudad)

    if output_file is not None:
        output_path = Path(output_file)
    else:
        target_dir = output_dir or OUTPUT_DIR
        target_dir.mkdir(parents=True, exist_ok=True)
        output_path = build_output_path(category_slug, ciudad_slug, target_dir)

    if leads:
        if verbose:
            print("\n[3/3] Exportando {} negocios ({} con email)...".format(len(leads), emails_found))
        export_to_excel(leads, output_path)
        if verbose:
            print("\nListo -> {}".format(output_path))

    if progress:
        progress("Listo! Preparando descarga...", 1, 1)

    return ScrapeResult(
        filename=output_path.name,
        path=output_path,
        count=len(leads),
        rubro=category.label,
        ciudad=ciudad,
        emails_found=emails_found,
        total_reported=len(leads),
    )


def scrape_leads_multi_city(category, cities, output_dir=None, verbose=True, fetch_emails=True, progress=None):
    if not cities:
        raise ValueError("Ingresa al menos una ciudad.")
    if len(cities) > MAX_CITIES:
        raise ValueError("Maximo {} ciudades por busqueda.".format(MAX_CITIES))
    check_api_key()

    all_leads    = []
    n_queries    = len(category.queries)
    total_search = n_queries * len(cities)

    for idx, ciudad in enumerate(cities, start=1):
        if verbose:
            print("\n[Ciudad {}/{}] {}".format(idx, len(cities), ciudad))
        city_leads = search_places_for_city(
            category, ciudad.strip(),
            progress=progress,
            query_offset=(idx - 1) * n_queries,
            total_queries=total_search,
        )
        all_leads.extend(city_leads)

    all_leads = deduplicate(all_leads)
    with_web  = sum(1 for l in all_leads if l.get("sitio_web"))

    if fetch_emails:
        if verbose:
            print("\n[Email] Buscando emails en {} sitios...".format(with_web))
        all_leads = enrich_with_emails(
            all_leads,
            progress=progress,
            email_offset=total_search,
            total_emails=total_search + with_web,
        )

    emails_found  = sum(1 for l in all_leads if l.get("email"))
    category_slug = slugify(category.label)
    cities_slug   = "_".join(slugify(c) for c in cities[:3])
    if len(cities) > 3:
        cities_slug += "_y{}mas".format(len(cities) - 3)

    target_dir  = output_dir or OUTPUT_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = target_dir / "leads_{}_{}_{}.xlsx".format(category_slug, cities_slug, ts)

    if all_leads:
        if verbose:
            print("\n[Excel] Exportando {} negocios unicos ({} con email)...".format(len(all_leads), emails_found))
        export_to_excel(all_leads, output_path)
        if verbose:
            print("\nListo -> {}".format(output_path))

    if progress:
        progress("Listo! Preparando descarga...", 1, 1)

    return ScrapeResult(
        filename=output_path.name,
        path=output_path,
        count=len(all_leads),
        rubro=category.label,
        ciudad=", ".join(cities),
        emails_found=emails_found,
        total_reported=len(all_leads),
    )


def scrape_leads_by_index(category_index, ciudad, **kwargs):
    return scrape_leads(get_category(category_index), ciudad, **kwargs)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def print_category_menu():
    print("\n--- Categorias ---\n")
    for i, cat in enumerate(CATEGORIES, start=1):
        print("  {:2}. {}".format(i, cat.label))


def prompt_category():
    print_category_menu()
    while True:
        choice = input("\nElegi una categoria (1-{}): ".format(len(CATEGORIES))).strip()
        if not choice.isdigit():
            print("Ingresa el numero.")
            continue
        try:
            return get_category(int(choice))
        except ValueError as exc:
            print(exc)


def print_city_menu():
    print("\n--- Ciudades ({}) ---\n".format(PROVINCE))
    for i, city in enumerate(CITIES, start=1):
        print("  {:2}. {}".format(i, city))


def prompt_cities():
    print_city_menu()
    print("\nPodes elegir hasta {} ciudades separadas por coma (ej: 1,3,5)".format(MAX_CITIES))
    while True:
        raw   = input("Elegi ciudades: ").strip()
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        if not parts:
            print("Ingresa al menos un numero.")
            continue
        if len(parts) > MAX_CITIES:
            print("Maximo {} ciudades.".format(MAX_CITIES))
            continue
        cities, valid = [], True
        for p in parts:
            if not p.isdigit():
                print("'{}' no es valido.".format(p))
                valid = False
                break
            try:
                cities.append(get_city(int(p)))
            except ValueError as exc:
                print(exc)
                valid = False
                break
        if valid and cities:
            return cities


def cli_progress(msg, current, total):
    if total > 0:
        pct = int((current / total) * 100)
        print("  [{:3d}%] {}".format(pct, msg))
    else:
        print("  {}".format(msg))


def main():
    print("\n--- PiP Studio - Buscador de Leads ---")
    category = prompt_category()
    cities   = prompt_cities()
    try:
        if len(cities) == 1:
            result = scrape_leads(category, cities[0], verbose=True, progress=cli_progress)
        else:
            result = scrape_leads_multi_city(category, cities, verbose=True, progress=cli_progress)
        print("\nResultado: {} negocios, {} con email.".format(result.count, result.emails_found))
    except requests.RequestException as exc:
        print("\nError de API: {}".format(exc))
        sys.exit(1)
    except ValueError as exc:
        print("Error: {}".format(exc))
        sys.exit(1)
    if result.count == 0:
        print("\nSin resultados.")
        sys.exit(0)


if __name__ == "__main__":
    main()

# ---------------------------------------------------------------------------
# Instalacion:
#   python -m pip install requests openpyxl beautifulsoup4
#
# Ejecucion CLI:   python scraper.py
# Servidor web:    python server.py
# ---------------------------------------------------------------------------